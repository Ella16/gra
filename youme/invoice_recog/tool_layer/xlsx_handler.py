import pandas as pd
from openpyxl import load_workbook

from invoice_recog.utils import utils

from . import logger


class xlsxHandler(object):
    def __init__(self, config: dict):
        self.config = config["tool_layer"]["output_delivery"]
        try:
            self.xlsx_file_name: str | list[str] = self.config["xlsx_file_name"]
            self.xlsx_sheet: str = self.config["xlsx_sheet_name"]
            self.key_mapping: dict[str, str] | str = self.config["key_mapping"]
            self.target_keys: list[str] | str = self.config["target_keys"]
            self.xlsx_file = None
            self.pdf_files = []
            logger.debug(f"[xlsx_handler] Initialized xlsxHandler")
        except Exception as e:
            logger.error(f"[xlsx_handler] Error initializing xlsxHandler: {e}")

    def set_target_files(self, folder_path: str) -> tuple[list[str], str]:
        if folder_path:
            folder_path = (
                folder_path if folder_path.endswith("/") else folder_path + "/"
            )
            xlsx_file = utils.load_files(folder_path, ext=self.xlsx_file_name.lower())
            pdf_files = utils.load_files(folder_path, ext=".pdf")
        else:
            xlsx_file = []
            pdf_files = []

        # remove processed pdf files
        if pdf_files:
            logger.info(f"[xlsx_handler] Found {len(pdf_files)} PDF files")
            if xlsx_file:
                processed_pdf_files = pd.read_excel(xlsx_file[0])[
                    self.config["pdf_file_column_name"]
                ].tolist()
                pdf_files = self._remove_processed_pdf_files(
                    pdf_files, processed_pdf_files
                )
                logger.info(f"[xlsx_handler] {len(pdf_files)} PDF files to process")

            # create xlsx file if not exists
            if not xlsx_file:
                xlsx_file = [
                    folder_path
                    + f"YAS_Invoice_{utils.get_current_time_for_xlsxfile()}.xlsx"
                ]  # YAS_Invoice_20241114.xlsx
                self._create_xlsx_file(xlsx_file[0], self.xlsx_sheet, self.target_keys)

            # sheet name check
            sheets_in_xlsx = pd.ExcelFile(xlsx_file[0]).sheet_names
            if self.xlsx_sheet not in sheets_in_xlsx:
                self.xlsx_sheet = sheets_in_xlsx[0]

        self.pdf_files = pdf_files
        self.xlsx_file = xlsx_file[0] if xlsx_file else None

        return (self.pdf_files, self.xlsx_file)

    def _create_xlsx_file(
        self, xlsx_file: str, sheet_name: str, column_names: list[str]
    ) -> None:
        if isinstance(column_names, str):
            column_names = [c.strip() for c in column_names.split(",")]

        if not (isinstance(column_names, list) or isinstance(column_names, str)):
            logger.error(f"[xlsx_handler] Invalid column names: {column_names}")

        df = pd.DataFrame(columns=column_names)
        df.to_excel(
            xlsx_file, sheet_name=sheet_name, index=False, header=True, na_rep=""
        )
        logger.info(f"[xlsx_handler] Created xlsx file: {xlsx_file}")

    def _remove_processed_pdf_files(self, files: list[str], processed_files: list[str]):
        return [f for f in files if f.split("/")[-1] not in processed_files]

    def _get_the_next_no(self) -> tuple[pd.DataFrame, int]:
        if self.xlsx_file:
            existing_xlsx_data = pd.read_excel(
                self.xlsx_file, sheet_name=self.xlsx_sheet, header=0
            )
            if existing_xlsx_data.empty:
                start_no = 1
            else:
                try:
                    start_no = (
                        len(existing_xlsx_data) + 1
                    )  # 제목 행 빼고 행수를 셌고, 다음 번호니깐 1 더함
                except:
                    start_no = 1001
        else:
            existing_xlsx_data = None
            start_no = 1

        return existing_xlsx_data, start_no

    def _set_no_for_xlsx(self, data: list[dict[str, any]] = [], start_no: int = 1):
        for d in data:
            d[self.config["number_column_name"]] = start_no
            d[self.config["worker_column_name"]] = "Growdle"
            start_no += 1
        return data

    ################################
    ### main function
    ### save result to xlsx file
    def run(self, result: dict[str, any]) -> None:
        # convert result to xlsx data format
        new_result = [utils.flatten_dict(v) for v in result.values()]
        converted_result = [self._convert_key_for_xlsx(v) for v in new_result]
        logger.debug(f"[xlsx_handler] Converted for columns of xlsx format")

        # start number for No. update. 엑셀 데이터에 이어서. from YAS.xlsx
        _, start_no = self._get_the_next_no()
        converted_result = self._set_no_for_xlsx(converted_result, start_no)
        df_for_xlsx = self._convert_list_for_xlsx(converted_result)
        logger.info(f"[xlsx_handler]\n{df_for_xlsx.head()}")
        logger.debug(f"[xlsx_handler] Set No. for xlsx data")

        # save to xlsx
        try:
            self._update_xlsx(
                new_data=df_for_xlsx,
                file_path=self.xlsx_file,
                sheet_name=self.xlsx_sheet,
            )
        except Exception as e:
            logger.error(
                f"[xlsx_handler] Error saving result to xlsx file - {self.xlsx_file}"
            )
            logger.error(f"[xlsx_handler] {e}")

        # set cell width
        try:
            self._set_cell_width()
        except:
            logger.error(f"[xlsx_handler] Error setting cell width")

    def _set_cell_width(self):
        wb = load_workbook(self.xlsx_file)
        ws = wb[self.xlsx_sheet]

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (min(max_length, 61) + 2) * 1.5
            ws.column_dimensions[column].width = adjusted_width

        wb.save(self.xlsx_file)

    def _update_xlsx(
        self, new_data: pd.DataFrame, file_path: str, sheet_name: str
    ) -> bool:
        try:
            with pd.ExcelWriter(
                file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
            ) as writer:
                new_data.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=False,
                    na_rep="",
                    startrow=writer.sheets[sheet_name].max_row,
                )
            logger.info(f"[utils] Update {len(new_data)} rows to {file_path}")
            return True
        except Exception as e:
            logger.error(f"[utils] Failed to update xlsx to {file_path}: {e}")
            return False

    def _convert_key_for_xlsx(self, data: dict[str, any]) -> pd.DataFrame:
        # data = {
        #     "contents": "```markdown```",
        #     "invoice_id": "U488077/0007",
        #     "currency": "USD",
        #     "associated_text": "Unique Invoice ID, Cost, US$ 1,584.15",
        #     "total_amount": 1584.15,
        #     "pdf_file_name": "OPP20103666REPDN_240313(16)_@ANYM648643.PDF",
        #     "start_timestamp": "11/13/2024 16:32:47",
        #     "cost": (0.001913, 0.00395, 0.00334),
        #     "end_timestamp": "11/13/2024 16:32:59",
        # }
        # target_key = [No.,원본파일명,검색시작시간,검색종료시간,작업자,외국청구번호,총비용,화폐,GW접수자,GW접수번호,GW저장시간,GW저장파일명]
        # key_mapping = {invoice_id:외국청구번호,currency:화폐,total_amount:총비용,pdf_file_name:원본파일명,start_timestamp:검색시작시간,end_timestamp:검색종료시간}
        key_mapping = self.key_mapping
        target_keys = self.target_keys

        if isinstance(key_mapping, str):
            key_mapping = {
                k.split(":")[0].strip(): k.split(":")[1].strip()
                for k in key_mapping.split(",")
            }
        key_mapping_reverse = {v: k for k, v in key_mapping.items()}
        self.key_mapping = key_mapping

        if isinstance(target_keys, str):
            target_keys = [t.strip() for t in target_keys.split(",")]

        try:
            # target dictionary for xlsx
            response = {}
            for target_key in target_keys:
                response[target_key] = (
                    data[key_mapping_reverse[target_key]]
                    if target_key in key_mapping_reverse
                    else ""
                )
            return response
        except Exception as e:
            logger.error(f"[utils] Error converting dict to df: {e}")
            return {}

    def _convert_list_for_xlsx(self, data: list[dict[str, any]] = []) -> pd.DataFrame:
        return pd.DataFrame(data)
